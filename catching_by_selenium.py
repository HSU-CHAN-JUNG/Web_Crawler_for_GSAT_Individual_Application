# 載入 selenium4 的相關模組 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
# from webdriver_manager.chrome import ChromeDriverManager
# 載入 openpyxl 的相關模組
from openpyxl import Workbook
# 資料存取和時間的封包
import pandas as pd
import time
# 設定 chorme driver 的執行路徑

options = Options()
options.chrome_executable_path = r"C:\Users\User\python\chromedriver_win32\chromedriver.exe"
options.add_experimental_option('excludeSwitches', ['enable-logging'])

# 建立 Driver 物件實體，用程式操作瀏覽器運行
# driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
driver = webdriver.Chrome(options=options)
# 到校系分則的網站
# driver.get("https://www.cac.edu.tw/apply113/system/ColQry_vforStu113apply_GF84ad9zx/TotalGsdShow.htm")
driver.get("https://www.cac.edu.tw/apply115/system/ColQry_115xappLyfOrStu_Azd5gP29/TotalGsdShow.htm")

# 設定i來執行儲存新的單位格
i = 1
# 打開一個新的Excel及標籤頁
wb = Workbook()
ws = wb.active
# 抓取所有學校的連結，接著跑學校
universities = driver.find_elements(By.CSS_SELECTOR,"a")
for university in universities :
# university = driver.find_element(By.LINK_TEXT, "(016)國立中央大學 -38系(組)")
    university.click()
    # departments = driver.find_elements(By.LINK_TEXT, "詳細資料")
    departments = driver.find_elements(By.CSS_SELECTOR, "a")
    for department in departments:
        department.click()
        # 打開的PDF是新的分頁，所以要跳轉到新的分頁[-1]
        new_window = driver.window_handles
        driver.switch_to.window(new_window[-1])
        # 等個一秒，讓資料跑出來
        time.sleep(1)
        # 獲得學校系所和採計科目
        colloge = driver.find_element(By.CSS_SELECTOR, "div[class = 'colname']")
        department = driver.find_element(By.CSS_SELECTOR, "div[class = 'gsdname']")
        objects1 = driver.find_elements(By.CSS_SELECTOR, "td[class = 'Bb font_bold g3']")
        objects2 = driver.find_elements(By.CSS_SELECTOR, "td[colspan = '2']")
        objects3 = driver.find_elements(By.CSS_SELECTOR, "td[rowspan = '7']")
        
        '''
        objects1 = driver.find_elements(By.CSS_SELECTOR, "td[class = 'Bb font_bold']")
        objects2 = objects1.find_elements(By.CSS_SELECTOR, "td[colspan = '2']")
        objects = driver.find_elements(By.CSS_SELECTOR, "td[rowspan = '7']")
        '''
        
        objects = set(objects1) & set(objects2) & set(objects3)
        print(objects)
        # 最長等待1秒，每0.2秒檢查一次條件是否成立-->沒有成功
        # WebDriverWait(driver, 1, 0.2).until(EC.presence_of_elements_located("備註")) 

        # 轉換成字串格式
        for object in objects:
            str1 = str(object.text)
            str1 = str1.split('\n')
        # 學校系所寫入Excel
        ws.cell(i, 1).value = colloge.text 
        ws.cell(i, 2).value = department.text
        # 判斷採計科目，結果寫入Excel
        if ('數學A' in str1) & ('數學B' in str1):
            # print(colloge.text ,department.text ,"都採計")
            ws.cell(i, 3).value = "均採計"
        elif '數學A' in str1 :
            # print(colloge.text ,department.text ,"數學A")
            ws.cell(i, 3).value = "數學A"
        elif '數學B' in str1 :
            # print(colloge.text ,department.text ,"數學B")
            ws.cell(i, 3).value = "數學B"
        else : 
            # print(colloge.text ,department.text ,'都不採計')
            ws.cell(i, 3).value = "均不採計"
        # Excel存檔
        wb.save('115採計數學.xlsx') 
        i = i + 1
        # 關閉當前分頁，回到前一頁
        driver.close()
        driver.switch_to.window(new_window[0])
    # 回到前一頁
    driver.back()

driver.close()