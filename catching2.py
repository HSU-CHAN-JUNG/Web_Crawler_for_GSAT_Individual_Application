import os
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook

# 1. 設定初始網址與 Headers 偽裝
base_url = "https://www.cac.edu.tw/apply115/system/ColQry_115xappLyfOrStu_Azd5gP29/"
start_url = urljoin(base_url, "TotalGsdShow.htm")

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
}

# 2. 建立 Excel 活頁簿與標頭
wb = Workbook()
ws = wb.active
ws.append(["學校", "系所", "數學採計狀態"])  # 先寫入標頭，方便閱讀
excel_file = '採計數學.xlsx'

print("開始爬取 115 學年度個人申請校系分則...")

# 啟用 requests Session 提高連線效率
session = requests.Session()
session.headers.update(headers)

try:
    response = session.get(start_url)
    if response.status_code != 200:
        print(f"無法連線到主頁面，狀態碼：{response.status_code}")
        exit()
        
    # 解析主頁面（所有學校列表）
    soup = BeautifulSoup(response.content, "html.parser", from_encoding="utf-8")
    
    # 抓取所有學校的連結（過濾掉非學校簡章的 href，通常包含 'html' 或 'htm'）
    university_links = []
    for a in soup.find_all("a", href=True):
        # print(a)
        href = a["href"]
        # 排除無關連結，只拿學校代碼相關的網址
        # if "htm" in href and "TotalGsdShow" not in href:
        university_links.append(urljoin(base_url, href))
    

    # 移除重複的連結並保持順序
    university_links = list(dict.fromkeys(university_links))
    print(f"共找到 {len(university_links)} 所學校，開始深入抓取系所資料...")

    row_idx = 2  # 從第二行開始寫入資料（第一行是標頭）
    
    # 3. 遍歷每所學校
    for uni_url in university_links:
        print(uni_url)
        uni_response = session.get(uni_url)
        if uni_response.status_code != 200:
            continue
            
        # uni_soup = BeautifulSoup(uni_response.content, "html.parser", from_encoding="utf-8")
        uni_soup = BeautifulSoup(uni_response.content, "html5lib")
        
        # 抓取該學校內所有系所的詳細資料連結
        dept_links = []
        for a in uni_soup.find_all("a", href=True):
            print(a.text)
            if "詳細資料" in a.text or "./html/" in a["href"]:
                # 自動將 ./html/115_001012.htm?v=1.0 轉為完整網址
                full_url = urljoin(base_url, a["href"])
                dept_links.append(full_url)

        dept_links = list(dict.fromkeys(dept_links))
        
        print(dept_links)
        # 4. 遍歷每個系所的詳細頁面
        for dept_url in dept_links:
            try:
                dept_response = session.get(dept_url)
                if dept_response.status_code != 200:
                    continue
                    
                dept_soup = BeautifulSoup(dept_response.content, "html.parser", from_encoding="utf-8")
                
                # 取得學校名稱與系所名稱
                college_el = dept_soup.find("div", class_="colname")
                dept_el = dept_soup.find("div", class_="gsdname")
                
                print(college_el, dept_el)
                break

                if not college_el or not dept_el:
                    continue
                    
                college_text = college_el.text.strip()
                dept_text = dept_el.text.strip()
                
                # 5. 修改原本 Selenium 聯集/交集不穩定的問題
                # 直接撈取整個網頁的文字，或者撈取表格(td)內的文字來判斷最精準
                page_text = dept_soup.get_text()
                
                # 判斷採計科目
                has_mathA = "數學A" in page_text
                has_mathB = "數學B" in page_text
                
                if has_mathA and has_mathB:
                    math_status = "均採計"
                elif has_mathA:
                    math_status = "數學A"
                elif has_mathB:
                    math_status = "數學B"
                else:
                    math_status = "均不採計"
                
                # 寫入 Excel
                ws.cell(row_idx, 1).value = college_text
                ws.cell(row_idx, 2).value = dept_text
                ws.cell(row_idx, 3).value = math_status
                
                print(f"已記錄：{college_text} - {dept_text} ({math_status})")
                row_idx += 1
                
            except Exception as e:
                print(f"處理系所頁面時發生錯誤: {dept_url}, 錯誤: {e}")
                
        # 爬完一所學校就存檔一次，避免程式中斷資料遺失
        wb.save(excel_file)

    print(f"\n🎉 爬取完成！資料已成功儲存至 {excel_file}")

except Exception as e:
    print(f"程式執行發生異常: {e}")